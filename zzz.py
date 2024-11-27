import sys

import os

from concurrent.futures import ThreadPoolExecutor

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,

                              QHBoxLayout, QPushButton, QLabel, QFileDialog,

                              QTableWidget, QTableWidgetItem, QProgressBar,

                              QTextEdit, QStyle, QHeaderView, QMessageBox)

from PySide6.QtCore import Qt, QThread, Signal

from PySide6.QtGui import QColor, QIcon, QFont

import pandas as pd

import pdfplumber  # Mejor extracción de texto que PyPDF2

import sys

import os

from concurrent.futures import ThreadPoolExecutor

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,

                              QHBoxLayout, QPushButton, QLabel, QFileDialog,

                              QTableWidget, QTableWidgetItem, QProgressBar,

                              QTextEdit, QStyle, QHeaderView, QMessageBox)

from PySide6.QtCore import Qt, QThread, Signal

from PySide6.QtGui import QColor, QIcon, QFont

import pandas as pd

import pdfplumber  # Mejor extracción de texto que PyPDF2

import re

from datetime import datetime

import fitz

import pandas as pd

from openpyxl.styles import PatternFill

from openpyxl import load_workbook

from openpyxl.styles import Font, Alignment

from openpyxl.utils import get_column_letter





class WorkerThread(QThread):

    progress = Signal(int)

    file_processed = Signal(dict)

    error_occurred = Signal(str)

   

    def __init__(self, folder_path):

        super().__init__()

        self.folder_path = folder_path

       

    def get_pdf_files(self):

        """Obtiene todos los archivos PDF de la carpeta y subcarpetas"""

        pdf_files = []

        for root, dirs, files in os.walk(self.folder_path):

            for file in files:

                if file.lower().endswith('.pdf'):

                    pdf_files.append(os.path.join(root, file))

        return pdf_files

 

    import re

 

    import re

 

    def extraer_datos_texto(self, texto):

        datos = {}

 

        patrones = {

            'num_resolucion': r'(?i)resoluci[oó]n\s+n[úu]mero\s+(\b(?:uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|once|doce|trece|catorce|quince|diecis[eé]is|diecisiete|dieciocho|diecinueve|veinte|veintiuno|veintid[oó]s|veintitr[eé]s|veinticuatro|veinticinco|veintis[eé]is|veintisiete|veintiocho|veintinueve|treinta|treinta y uno|\d+)\b)',

            'fecha_resolucion_incompleta': r'(?i)(\d{1,2}\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+del?\s+202\d? )',

            'fecha_resolucion_larga': r'(?i)\b(?:uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|once|doce|trece|catorce|quince|diecis[eé]is|diecisiete|dieciocho|diecinueve|veinte|veintiuno|veintid[oó]s|veintitr[eé]s|veinticuatro|veinticinco|veintis[eé]is|veintisiete|veintiocho|veintinueve|treinta|treinta y uno)\b\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+del\s+año\s+(dos\s+mil(?:\s+(?:uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez))?)',

            'fecha_resolucion_numerica': r'(?i)(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+del?\s+(\d{4})',

           

            'tipo_materia': r'(?i)medida\s*[:\s]*\s*(.*?)\s*(?=\n|$)',

            'tipo_documento': r'(?i)(DNI|RUC)',

            'nro_documento': r'(?i)(DNI|RUC)\s*(?:N[°o"\.]*\s*)?\s*([0-9]{8})',

            'juzgado': r'(?i)\b(?:[0-9]+\'?\s*)?JUZGADO\s+DE\s+PAZ\s+LETRADO(?:\s+[A-Z]+)?\s+-?\s+[A-Z\s]+',

            'nombre':r'([A-ZÁÉÍÓÚÑ\s]+)\s+(?:con\s+N[\'"]?\s*de\s*DNI|identificado\s+con\s+DNI)',

            'num_exp': r'(?i)(?:expediente\s*[:\s]+)?([0-9]{5}-[0-9]{4}-[0-9]-[A-Z0-9\s-]+)',

            'direccion_juzgado': '',  # A definir patrón de regex

            'nombre_secretario_especialista': r'(?i)ESPECIALISTA\s+([A-Z\s]+)'

        }

 

        # Limitar la búsqueda a las primeras líneas del texto

        primeras_lineas = '\n'.join(texto.split('\n')[:50])

        #print(primeras_lineas)

        for campo, patron in patrones.items():

            if campo in ['fecha_resolucion_incompleta','fecha_resolucion_numerica','fecha_resolucion_larga']:

                # Buscar primero en las primeras líneas

                match = re.search(patron, primeras_lineas, re.MULTILINE | re.DOTALL)

                if not match:

                    # Si no se encuentra, buscar en todo el texto

                    match = re.search(patron, texto, re.MULTILINE | re.DOTALL)

                if match and 'fecha_resolucion' not in datos:

                    datos['fecha_resolucion'] = match.group(0).strip()

            else:

                if patron:  # Solo buscar si el patrón está definido

                    match = re.search(patron, texto, re.MULTILINE | re.DOTALL)

                    if match:

                        if campo == 'tipo_documento':

                            datos[campo] = match.group(1)

                        elif campo == 'nro_documento':

                            # Usar la misma coincidencia para obtener el número de documento asociado

                            numero_match = re.search(r'(?i)(DNI|RUC)\s*(?:N[°o"\.]*\s*)?\s*([0-9]{8})', match.group(0))

                            datos[campo] = numero_match.group(2) if numero_match else ''

                        else:

                            datos[campo] = match.group(1).strip() if campo in ['fecha_resolucion_larga', 'fecha_resolucion_numerica', 'fecha_resolucion_incompleta'] else match.group(0).strip()

                    else:

                        datos[campo] = ''  # Valor vacío si no se encuentra coincidencia

                else:

                    datos[campo] = ''  # Valor vacío si el patrón no está definido

       

        return datos




   

    def leer_pdf(self, ruta):

        texto_completo = ""

        try:

            documento = fitz.open(ruta)

            for num_pagina in range(documento.page_count):

                pagina = documento.load_page(num_pagina)

                texto_completo += pagina.get_text("text")

            documento.close()

        except Exception as e:

            raise Exception(f"Error leyendo PDF: {str(e)}")

        return texto_completo

   

    def run(self):

        pdf_files = self.get_pdf_files()

        total_files = len(pdf_files)

       

        if total_files == 0:

            self.error_occurred.emit("No se encontraron archivos PDF en la carpeta")

            return

           

        for i, archivo in enumerate(pdf_files, 1):

            try:

                contenido = self.leer_pdf(archivo)

                if contenido:

                    datos = self.extraer_datos_texto(contenido)

                    datos['archivo'] = os.path.basename(archivo)

                    datos['ruta'] = archivo

                    self.file_processed.emit(datos)

                else:

                    self.error_occurred.emit(f"No se pudo extraer texto de: {archivo}")

            except Exception as e:

                self.error_occurred.emit(f"Error procesando {archivo}: {str(e)}")

           

            self.progress.emit(int((i / total_files) * 100))




class MainWindow(QMainWindow):

    def __init__(self):

        super().__init__()

        self.setWindowTitle("Extractor de Datos de Oficios PDF")

        self.setMinimumSize(1200, 800)

        self.datos_extraidos = []

       

        # Configurar la interfaz

        self.setup_ui()

       

    def setup_ui(self):

        # Widget y layout principal

        main_widget = QWidget()

        self.setCentralWidget(main_widget)

        layout = QVBoxLayout(main_widget)

       

        # Estilo general

        self.setStyleSheet("""

            QMainWindow {

                background-color: #f0f2f5;

            }

            QPushButton {

                background-color: #1976D2;

                color: white;

                border: none;

                padding: 8px 16px;

                border-radius: 4px;

                font-weight: bold;

                min-width: 150px;

            }

            QPushButton:hover {

                background-color: #1565C0;

            }

            QPushButton:pressed {

                background-color: #0D47A1;

            }

            QTableWidget {

                background-color: white;

                border: 1px solid #ddd;

                border-radius: 4px;

                gridline-color: #ddd;

            }

            QHeaderView::section {

                background-color: #f5f5f5;

                padding: 6px;

                border: 1px solid #ddd;

                font-weight: bold;

            }

            QProgressBar {

                border: 1px solid #ddd;

                border-radius: 4px;

                text-align: center;

                height: 20px;

            }

            QProgressBar::chunk {

                background-color: #1976D2;

            }

            QTextEdit {

                background-color: white;

                border: 1px solid #ddd;

                border-radius: 4px;

                padding: 5px;

            }

        """)

       

        # Sección superior

        top_layout = QHBoxLayout()

       

        # Botones con íconos

        self.btn_seleccionar = QPushButton("Seleccionar Carpeta")

        self.btn_seleccionar.setIcon(self.style().standardIcon(QStyle.SP_DirIcon))

        self.btn_procesar = QPushButton("Procesar PDFs")

        self.btn_procesar.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))

        self.btn_excel = QPushButton("Generar Excel")

        self.btn_excel.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))

       

        top_layout.addWidget(self.btn_seleccionar)

        top_layout.addWidget(self.btn_procesar)

        top_layout.addWidget(self.btn_excel)

        top_layout.addStretch()

        layout.addLayout(top_layout)

       

        # Etiqueta de carpeta seleccionada

        self.lbl_carpeta = QLabel("Carpeta seleccionada: Ninguna")

        self.lbl_carpeta.setStyleSheet("padding: 5px; color: #666;")

        layout.addWidget(self.lbl_carpeta)

       

        # Barra de progreso

        self.progress_bar = QProgressBar()

        self.progress_bar.setTextVisible(True)

        layout.addWidget(self.progress_bar)

       

        # Tabla de resultados

        self.table = QTableWidget()

        self.table.setColumnCount(11)

        self.table.setHorizontalHeaderLabels([

            "Archivo", "Numero de resolucion", "Fecha resolucion", "Tipo materia",

            "Tipo de documento", "Numero de documento", "Juzgado", "Nombre","Num Exp", "Direccion juzgado",

            "nom_secretario y/o Especialista",

        ])

        header = self.table.horizontalHeader()

        header.setSectionResizeMode(QHeaderView.Stretch)

        layout.addWidget(self.table)

       

        # Log de procesamiento

        log_layout = QVBoxLayout()

        log_label = QLabel("Log de procesamiento:")

        log_label.setFont(QFont("Arial", 10, QFont.Bold))

        self.log_text = QTextEdit()

        self.log_text.setReadOnly(True)

        self.log_text.setMaximumHeight(150)

        log_layout.addWidget(log_label)

        log_layout.addWidget(self.log_text)

        layout.addLayout(log_layout)

       

        # Conectar señales

        self.btn_seleccionar.clicked.connect(self.seleccionar_carpeta)

        self.btn_procesar.clicked.connect(self.procesar_archivos)

        self.btn_excel.clicked.connect(self.generar_excel)

       

        self.carpeta_seleccionada = None

       

    def seleccionar_carpeta(self):

        carpeta = QFileDialog.getExistingDirectory(

            self,

            "Seleccionar Carpeta con PDFs",

            "",

            QFileDialog.ShowDirsOnly

        )

        if carpeta:

            self.carpeta_seleccionada = carpeta

            self.lbl_carpeta.setText(f"Carpeta seleccionada: {carpeta}")

            self.log_text.append(f"Carpeta seleccionada: {carpeta}")

            self.progress_bar.setValue(0)

   

    def procesar_archivos(self):

        if not self.carpeta_seleccionada:

            QMessageBox.warning(self, "Advertencia", "Por favor seleccione una carpeta primero")

            return

       

        self.datos_extraidos = []

        self.table.setRowCount(0)

        self.progress_bar.setValue(0)

       

        # Crear y configurar el worker thread

        self.worker = WorkerThread(self.carpeta_seleccionada)

        self.worker.progress.connect(self.actualizar_progreso)

        self.worker.file_processed.connect(self.agregar_datos_tabla)

        self.worker.error_occurred.connect(self.log_text.append)

        self.worker.start()

   

    def actualizar_progreso(self, valor):

        self.progress_bar.setValue(valor)

   

    def agregar_datos_tabla(self, datos):

        print("Datos a agregar:", datos)

        self.datos_extraidos.append(datos)

        row = self.table.rowCount()

        self.table.insertRow(row)

       

        campos = [

            'archivo', 'num_resolucion', 'fecha_resolucion', 'tipo_materia',

            'tipo_documento', 'nro_documento', 'juzgado', 'nombre','num_exp',

            'direccion_juzgado', 'nombre_secretario_especialista'

        ]

       

        for col, campo in enumerate(campos):

            item = QTableWidgetItem(datos.get(campo, ''))

            item.setTextAlignment(Qt.AlignCenter)

            # Hacer el item editable para correcciones manuales

            item.setFlags(item.flags() | Qt.ItemIsEditable)

            self.table.setItem(row, col, item)

   

    def generar_excel(self):

        if not self.datos_extraidos:

            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")

            return

           

        try:

            nombre_archivo, _ = QFileDialog.getSaveFileName(

                self,

                "Guardar Excel",

                "",

                "Excel (*.xlsx)"

            )

           

            if nombre_archivo:

                # Obtener datos actualizados de la tabla

                datos_actualizados = []

                for row in range(self.table.rowCount()):

                    fila = {}

                    for col in range(self.table.columnCount()):

                        header = self.table.horizontalHeaderItem(col).text()

                        item = self.table.item(row, col)

                        fila[header] = item.text() if item else ''

                    datos_actualizados.append(fila)

               

                # Crear el DataFrame y guardar en un archivo Excel

                df = pd.DataFrame(datos_actualizados)

                df.to_excel(nombre_archivo, index=False)

 

                # Abrir el archivo Excel para aplicar formato

                wb = load_workbook(nombre_archivo)

                ws = wb.active

 

                # Aplicar formato al encabezado

                for cell in ws["1:1"]:

                    cell.font = Font(bold=True, color="FFFFFF")

                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

 

                # Ajustar el tamaño de las columnas

                for col in ws.columns:

                    max_length = 0

                    column = col[0].column_letter  # Obtener la letra de la columna

                    for cell in col:

                        try:

                            if len(str(cell.value)) > max_length:

                                max_length = len(cell.value)

                        except:

                            pass

                    adjusted_width = (max_length + 2)

                    ws.column_dimensions[column].width = adjusted_width

 

                wb.save(nombre_archivo)

 

                self.log_text.append(f"Reporte Excel generado: {nombre_archivo}")

                QMessageBox.information(self, "Éxito", "Reporte generado exitosamente")

 

        except Exception as e:

            self.log_text.append(f"Error generando Excel: {str(e)}")

            QMessageBox.critical(self, "Error", f"Error al generar el archivo: {str(e)}")

 

def main():

    app = QApplication(sys.argv)

    window = MainWindow()

    window.show()

    sys.exit(app.exec())

 

if __name__ == "__main__":

    main()